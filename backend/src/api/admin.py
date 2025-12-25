"""
Admin API endpoints for Hub.
Includes user management, groups, access control, and admin tools.
"""
from datetime import datetime, timezone
from io import BytesIO
from typing import List, Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func, delete, and_, or_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from ..core.dependencies import get_current_admin
from ..db.base import get_db
from ..models import User, Application, UserGroup, ApplicationAccess, OAuthToken, OAuthCode
from ..models.user_group import user_group_members
from ..schemas.admin import (
    UserGroupCreate,
    UserGroupUpdate,
    UserGroupResponse,
    UserGroupListResponse,
    UserGroupMemberInfo,
    AccessGrantRequest,
    AccessRevokeRequest,
    ApplicationAccessResponse,
    AdminStatsResponse,
    UserListResponse,
    UserUpdateRequest,
    BulkUserActionRequest,
    BulkGroupMembershipRequest,
)

router = APIRouter(prefix="/api/admin", tags=["admin"])


# ============ User Management ============

@router.get("/users", response_model=List[UserListResponse])
async def list_users(
    search: Optional[str] = Query(None, description="Search by email or name"),
    is_admin: Optional[bool] = Query(None),
    is_active: Optional[bool] = Query(None),
    group_id: Optional[UUID] = Query(None, description="Filter by group"),
    limit: int = Query(100, le=500),
    offset: int = Query(0, ge=0),
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """List all users with filtering and search."""
    query = select(User).options(selectinload(User.groups))

    # Apply filters
    if search:
        search_pattern = f"%{search}%"
        query = query.where(
            or_(
                User.email.ilike(search_pattern),
                User.display_name.ilike(search_pattern),
                User.first_name.ilike(search_pattern),
                User.last_name.ilike(search_pattern),
            )
        )

    if is_admin is not None:
        query = query.where(User.is_admin == is_admin)

    if is_active is not None:
        query = query.where(User.is_active == is_active)

    if group_id:
        query = query.join(user_group_members).where(
            user_group_members.c.group_id == group_id
        )

    query = query.order_by(User.created_at.desc()).offset(offset).limit(limit)

    result = await db.execute(query)
    users = result.scalars().all()

    # Build response with group names and app counts
    response = []
    for user in users:
        # Count accessible apps
        access_count = await db.execute(
            select(func.count(ApplicationAccess.id)).where(
                ApplicationAccess.user_id == user.id
            )
        )
        direct_count = access_count.scalar() or 0

        response.append(UserListResponse(
            id=user.id,
            email=user.email,
            display_name=user.display_name,
            first_name=user.first_name,
            last_name=user.last_name,
            middle_name=user.middle_name,
            department=user.department,
            job_title=user.job_title,
            is_active=user.is_active,
            is_admin=user.is_admin,
            last_login_at=user.last_login_at,
            created_at=user.created_at,
            groups=[g.name for g in user.groups],
            app_count=direct_count,
        ))

    return response


@router.patch("/users/{user_id}", response_model=UserListResponse)
async def update_user(
    user_id: UUID,
    data: UserUpdateRequest,
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Update user properties (active, admin status)."""
    result = await db.execute(
        select(User).options(selectinload(User.groups)).where(User.id == user_id)
    )
    user = result.scalar_one_or_none()

    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # Prevent self-demotion
    if user.id == admin.id and data.is_admin is False:
        raise HTTPException(status_code=400, detail="Cannot remove your own admin status")

    if data.is_active is not None:
        user.is_active = data.is_active
    if data.is_admin is not None:
        user.is_admin = data.is_admin

    await db.commit()
    await db.refresh(user)

    return UserListResponse(
        id=user.id,
        email=user.email,
        display_name=user.display_name,
        first_name=user.first_name,
        last_name=user.last_name,
        middle_name=user.middle_name,
        department=user.department,
        job_title=user.job_title,
        is_active=user.is_active,
        is_admin=user.is_admin,
        last_login_at=user.last_login_at,
        created_at=user.created_at,
        groups=[g.name for g in user.groups],
        app_count=0,
    )


@router.post("/users/bulk")
async def bulk_user_action(
    data: BulkUserActionRequest,
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Perform bulk actions on users."""
    if admin.id in data.user_ids and data.action in ['deactivate', 'remove_admin']:
        raise HTTPException(status_code=400, detail="Cannot perform this action on yourself")

    result = await db.execute(
        select(User).where(User.id.in_(data.user_ids))
    )
    users = result.scalars().all()

    updated = 0
    for user in users:
        if data.action == 'activate':
            user.is_active = True
            updated += 1
        elif data.action == 'deactivate':
            user.is_active = False
            updated += 1
        elif data.action == 'make_admin':
            user.is_admin = True
            updated += 1
        elif data.action == 'remove_admin':
            if user.id != admin.id:
                user.is_admin = False
                updated += 1

    await db.commit()
    return {"updated": updated, "action": data.action}


# ============ Group Management ============

@router.get("/groups", response_model=List[UserGroupListResponse])
async def list_groups(
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """List all user groups."""
    result = await db.execute(
        select(UserGroup).options(selectinload(UserGroup.members)).order_by(UserGroup.name)
    )
    groups = result.scalars().all()

    return [
        UserGroupListResponse(
            id=g.id,
            name=g.name,
            description=g.description,
            color=g.color,
            member_count=len(g.members),
            created_at=g.created_at,
        )
        for g in groups
    ]


@router.post("/groups", response_model=UserGroupResponse, status_code=201)
async def create_group(
    data: UserGroupCreate,
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Create a new user group."""
    # Check name uniqueness
    existing = await db.execute(
        select(UserGroup).where(UserGroup.name == data.name)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Group name already exists")

    group = UserGroup(
        name=data.name,
        description=data.description,
        color=data.color,
        created_by=admin.id,
    )
    db.add(group)
    await db.flush()

    # Add initial members
    if data.member_ids:
        users_result = await db.execute(
            select(User).where(User.id.in_(data.member_ids))
        )
        users = users_result.scalars().all()
        group.members = list(users)

    await db.commit()
    await db.refresh(group)

    return UserGroupResponse(
        id=group.id,
        name=group.name,
        description=group.description,
        color=group.color,
        created_at=group.created_at,
        updated_at=group.updated_at,
        member_count=len(group.members),
        members=[
            UserGroupMemberInfo(
                id=m.id,
                email=m.email,
                display_name=m.display_name,
                first_name=m.first_name,
                last_name=m.last_name,
                department=m.department,
            )
            for m in group.members
        ],
    )


@router.get("/groups/{group_id}", response_model=UserGroupResponse)
async def get_group(
    group_id: UUID,
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Get group details with members."""
    result = await db.execute(
        select(UserGroup).options(selectinload(UserGroup.members)).where(UserGroup.id == group_id)
    )
    group = result.scalar_one_or_none()

    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    return UserGroupResponse(
        id=group.id,
        name=group.name,
        description=group.description,
        color=group.color,
        created_at=group.created_at,
        updated_at=group.updated_at,
        member_count=len(group.members),
        members=[
            UserGroupMemberInfo(
                id=m.id,
                email=m.email,
                display_name=m.display_name,
                first_name=m.first_name,
                last_name=m.last_name,
                department=m.department,
            )
            for m in group.members
        ],
    )


@router.patch("/groups/{group_id}", response_model=UserGroupResponse)
async def update_group(
    group_id: UUID,
    data: UserGroupUpdate,
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Update group properties."""
    result = await db.execute(
        select(UserGroup).options(selectinload(UserGroup.members)).where(UserGroup.id == group_id)
    )
    group = result.scalar_one_or_none()

    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    if data.name is not None:
        # Check uniqueness
        existing = await db.execute(
            select(UserGroup).where(UserGroup.name == data.name, UserGroup.id != group_id)
        )
        if existing.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="Group name already exists")
        group.name = data.name

    if data.description is not None:
        group.description = data.description
    if data.color is not None:
        group.color = data.color

    await db.commit()
    await db.refresh(group)

    return UserGroupResponse(
        id=group.id,
        name=group.name,
        description=group.description,
        color=group.color,
        created_at=group.created_at,
        updated_at=group.updated_at,
        member_count=len(group.members),
        members=[
            UserGroupMemberInfo(
                id=m.id,
                email=m.email,
                display_name=m.display_name,
                first_name=m.first_name,
                last_name=m.last_name,
                department=m.department,
            )
            for m in group.members
        ],
    )


@router.delete("/groups/{group_id}", status_code=204)
async def delete_group(
    group_id: UUID,
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Delete a group."""
    result = await db.execute(
        select(UserGroup).where(UserGroup.id == group_id)
    )
    group = result.scalar_one_or_none()

    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    await db.delete(group)
    await db.commit()


@router.post("/groups/{group_id}/members")
async def add_group_members(
    group_id: UUID,
    data: BulkGroupMembershipRequest,
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Add or remove members from a group."""
    result = await db.execute(
        select(UserGroup).options(selectinload(UserGroup.members)).where(UserGroup.id == group_id)
    )
    group = result.scalar_one_or_none()

    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    users_result = await db.execute(
        select(User).where(User.id.in_(data.user_ids))
    )
    users = users_result.scalars().all()

    current_member_ids = {m.id for m in group.members}
    updated = 0

    if data.action == 'add':
        for user in users:
            if user.id not in current_member_ids:
                group.members.append(user)
                updated += 1
    elif data.action == 'remove':
        group.members = [m for m in group.members if m.id not in data.user_ids]
        updated = len(data.user_ids)

    await db.commit()
    return {"updated": updated, "action": data.action}


# ============ Access Control ============

@router.get("/applications/{app_id}/access", response_model=ApplicationAccessResponse)
async def get_application_access(
    app_id: UUID,
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Get access configuration for an application."""
    result = await db.execute(
        select(Application).where(Application.id == app_id)
    )
    app = result.scalar_one_or_none()

    if not app:
        raise HTTPException(status_code=404, detail="Application not found")

    # Get direct user access
    direct_access = await db.execute(
        select(ApplicationAccess).options(selectinload(ApplicationAccess.user)).where(
            ApplicationAccess.application_id == app_id,
            ApplicationAccess.user_id.isnot(None)
        )
    )
    direct_users = [
        UserGroupMemberInfo(
            id=a.user.id,
            email=a.user.email,
            display_name=a.user.display_name,
            first_name=a.user.first_name,
            last_name=a.user.last_name,
            department=a.user.department,
        )
        for a in direct_access.scalars().all()
    ]

    # Get group access
    group_access = await db.execute(
        select(ApplicationAccess).options(
            selectinload(ApplicationAccess.group).selectinload(UserGroup.members)
        ).where(
            ApplicationAccess.application_id == app_id,
            ApplicationAccess.group_id.isnot(None)
        )
    )
    groups = [
        UserGroupListResponse(
            id=a.group.id,
            name=a.group.name,
            description=a.group.description,
            color=a.group.color,
            member_count=len(a.group.members),
            created_at=a.group.created_at,
        )
        for a in group_access.scalars().all()
    ]

    return ApplicationAccessResponse(
        application_id=app.id,
        application_name=app.name,
        is_public=app.is_public or False,
        direct_users=direct_users,
        groups=groups,
    )


@router.post("/access/grant")
async def grant_access(
    data: AccessGrantRequest,
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Grant access to an application for users or groups."""
    # Verify application exists
    app_result = await db.execute(
        select(Application).where(Application.id == data.application_id)
    )
    if not app_result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Application not found")

    granted = 0

    # Grant to users
    for user_id in data.user_ids:
        # Check if already exists
        existing = await db.execute(
            select(ApplicationAccess).where(
                ApplicationAccess.user_id == user_id,
                ApplicationAccess.application_id == data.application_id
            )
        )
        if not existing.scalar_one_or_none():
            access = ApplicationAccess(
                user_id=user_id,
                application_id=data.application_id,
                granted_by=admin.id,
            )
            db.add(access)
            granted += 1

    # Grant to groups
    for group_id in data.group_ids:
        existing = await db.execute(
            select(ApplicationAccess).where(
                ApplicationAccess.group_id == group_id,
                ApplicationAccess.application_id == data.application_id
            )
        )
        if not existing.scalar_one_or_none():
            access = ApplicationAccess(
                group_id=group_id,
                application_id=data.application_id,
                granted_by=admin.id,
            )
            db.add(access)
            granted += 1

    await db.commit()
    return {"granted": granted}


@router.post("/access/revoke")
async def revoke_access(
    data: AccessRevokeRequest,
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Revoke access to an application from users or groups."""
    revoked = 0

    # Revoke from users
    if data.user_ids:
        result = await db.execute(
            delete(ApplicationAccess).where(
                ApplicationAccess.user_id.in_(data.user_ids),
                ApplicationAccess.application_id == data.application_id
            )
        )
        revoked += result.rowcount

    # Revoke from groups
    if data.group_ids:
        result = await db.execute(
            delete(ApplicationAccess).where(
                ApplicationAccess.group_id.in_(data.group_ids),
                ApplicationAccess.application_id == data.application_id
            )
        )
        revoked += result.rowcount

    await db.commit()
    return {"revoked": revoked}


@router.patch("/applications/{app_id}/public")
async def set_application_public(
    app_id: UUID,
    is_public: bool = Query(...),
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Set whether an application is public (visible to all users)."""
    result = await db.execute(
        select(Application).where(Application.id == app_id)
    )
    app = result.scalar_one_or_none()

    if not app:
        raise HTTPException(status_code=404, detail="Application not found")

    app.is_public = is_public
    await db.commit()

    return {"application_id": app_id, "is_public": is_public}


# ============ Admin Stats & Tools ============

@router.get("/stats", response_model=AdminStatsResponse)
async def get_admin_stats(
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Get dashboard statistics."""
    now = datetime.now(timezone.utc)

    # User stats
    users_total = await db.execute(select(func.count(User.id)))
    users_active = await db.execute(select(func.count(User.id)).where(User.is_active == True))
    users_admin = await db.execute(select(func.count(User.id)).where(User.is_admin == True))

    # App stats
    apps_total = await db.execute(select(func.count(Application.id)))
    apps_active = await db.execute(select(func.count(Application.id)).where(Application.is_active == True))

    # Group stats
    groups_total = await db.execute(select(func.count(UserGroup.id)))

    # Token stats
    tokens_total = await db.execute(select(func.count(OAuthToken.id)))
    tokens_active = await db.execute(
        select(func.count(OAuthToken.id)).where(
            OAuthToken.expires_at > now,
            OAuthToken.revoked_at.is_(None)
        )
    )

    return AdminStatsResponse(
        users={
            "total": users_total.scalar() or 0,
            "active": users_active.scalar() or 0,
            "admins": users_admin.scalar() or 0,
        },
        applications={
            "total": apps_total.scalar() or 0,
            "active": apps_active.scalar() or 0,
        },
        groups={
            "total": groups_total.scalar() or 0,
        },
        tokens={
            "total": tokens_total.scalar() or 0,
            "active": tokens_active.scalar() or 0,
        },
        database={
            "status": "healthy",
        },
        generated_at=now,
    )


@router.post("/cleanup-tokens")
async def cleanup_tokens(
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Clean up expired OAuth codes and tokens."""
    now = datetime.now(timezone.utc)

    # Delete expired codes
    codes_result = await db.execute(
        delete(OAuthCode).where(OAuthCode.expires_at < now)
    )

    # Delete expired tokens
    tokens_result = await db.execute(
        delete(OAuthToken).where(OAuthToken.expires_at < now)
    )

    await db.commit()

    return {
        "deleted_codes": codes_result.rowcount,
        "deleted_tokens": tokens_result.rowcount,
    }


@router.get("/export/users")
async def export_users_excel(
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Export all users to Excel file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl not installed. Run: pip install openpyxl"
        )

    result = await db.execute(
        select(User).options(selectinload(User.groups)).order_by(User.created_at.desc())
    )
    users = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Users"

    # Headers
    headers = ['Email', 'Display Name', 'First Name', 'Last Name', 'Middle Name',
               'Department', 'Job Title', 'Groups', 'Admin', 'Active', 'Last Login', 'Created']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Data
    for row, user in enumerate(users, 2):
        ws.cell(row=row, column=1, value=user.email)
        ws.cell(row=row, column=2, value=user.display_name)
        ws.cell(row=row, column=3, value=user.first_name)
        ws.cell(row=row, column=4, value=user.last_name)
        ws.cell(row=row, column=5, value=user.middle_name)
        ws.cell(row=row, column=6, value=user.department)
        ws.cell(row=row, column=7, value=user.job_title)
        ws.cell(row=row, column=8, value=', '.join(g.name for g in user.groups))
        ws.cell(row=row, column=9, value='Yes' if user.is_admin else 'No')
        ws.cell(row=row, column=10, value='Yes' if user.is_active else 'No')
        ws.cell(row=row, column=11, value=user.last_login_at.strftime('%Y-%m-%d %H:%M') if user.last_login_at else '')
        ws.cell(row=row, column=12, value=user.created_at.strftime('%Y-%m-%d %H:%M'))

    # Auto-width
    for column in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"hub_users_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/applications")
async def export_applications_excel(
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Export all applications to Excel file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl not installed. Run: pip install openpyxl"
        )

    result = await db.execute(
        select(Application).order_by(Application.name)
    )
    apps = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"

    headers = ['Name', 'Slug', 'Client ID', 'Base URL', 'Active', 'Public', 'Created']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for row, app in enumerate(apps, 2):
        ws.cell(row=row, column=1, value=app.name)
        ws.cell(row=row, column=2, value=app.slug)
        ws.cell(row=row, column=3, value=app.client_id)
        ws.cell(row=row, column=4, value=app.base_url)
        ws.cell(row=row, column=5, value='Yes' if app.is_active else 'No')
        ws.cell(row=row, column=6, value='Yes' if app.is_public else 'No')
        ws.cell(row=row, column=7, value=app.created_at.strftime('%Y-%m-%d %H:%M'))

    for column in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"hub_applications_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
