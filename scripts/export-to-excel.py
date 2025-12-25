#!/usr/bin/env python3
"""
Hub - Export Database to Excel
Exports users, applications, and tokens to Excel file.

Usage:
    python export-to-excel.py                    # Export all tables
    python export-to-excel.py --tables users     # Export only users
    python export-to-excel.py --output report.xlsx

Requirements:
    pip install openpyxl psycopg2-binary python-dotenv

Run from project root or specify DATABASE_URL environment variable.
"""

import os
import sys
import argparse
from datetime import datetime
from pathlib import Path

try:
    import psycopg2
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"Missing dependency: {e}")
    print("Install with: pip install openpyxl psycopg2-binary")
    sys.exit(1)

# Try to load .env
try:
    from dotenv import load_dotenv
    # Try multiple possible .env locations
    for env_file in ['.env', '.env.prod', '../.env', '../.env.prod']:
        if Path(env_file).exists():
            load_dotenv(env_file)
            break
except ImportError:
    pass


def get_db_connection():
    """Get database connection from environment or Docker."""
    db_url = os.getenv('DATABASE_URL')

    if db_url:
        # Parse DATABASE_URL
        # Format: postgresql+asyncpg://user:pass@host:port/db
        db_url = db_url.replace('postgresql+asyncpg://', 'postgresql://')
        return psycopg2.connect(db_url)

    # Default Docker connection
    return psycopg2.connect(
        host=os.getenv('DB_HOST', 'localhost'),
        port=os.getenv('DB_PORT', '5432'),
        database=os.getenv('DB_NAME', 'hub_db'),
        user=os.getenv('DB_USER', 'hub'),
        password=os.getenv('DB_PASSWORD', 'hubpassword')
    )


def style_header(cell):
    """Apply header styling."""
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


def style_cell(cell, row_idx):
    """Apply cell styling with alternating rows."""
    if row_idx % 2 == 0:
        cell.fill = PatternFill(start_color="E9EFF7", end_color="E9EFF7", fill_type="solid")
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.alignment = Alignment(vertical="center")


def auto_width(ws):
    """Auto-adjust column widths."""
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                cell_len = len(str(cell.value or ""))
                if cell_len > max_length:
                    max_length = min(cell_len, 50)  # Cap at 50
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2


def export_table(cursor, ws, table_name, query, headers):
    """Export a table to worksheet."""
    cursor.execute(query)
    rows = cursor.fetchall()

    # Headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        style_header(cell)

    # Data
    for row_idx, row in enumerate(rows, 2):
        for col, value in enumerate(row, 1):
            # Convert values
            if isinstance(value, (list, dict)):
                value = str(value)
            elif isinstance(value, datetime):
                value = value.strftime('%Y-%m-%d %H:%M:%S')

            cell = ws.cell(row=row_idx, column=col, value=value)
            style_cell(cell, row_idx)

    auto_width(ws)
    ws.freeze_panes = 'A2'  # Freeze header row

    return len(rows)


def main():
    parser = argparse.ArgumentParser(description='Export Hub database to Excel')
    parser.add_argument('--output', '-o', default=None,
                        help='Output file path (default: hub_export_YYYYMMDD.xlsx)')
    parser.add_argument('--tables', '-t', nargs='+',
                        choices=['users', 'applications', 'tokens', 'codes'],
                        default=['users', 'applications', 'tokens'],
                        help='Tables to export (default: users applications tokens)')
    args = parser.parse_args()

    # Default output filename
    if args.output is None:
        args.output = f"hub_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    print(f"=== Hub Database Export ===")
    print(f"Output: {args.output}")
    print(f"Tables: {', '.join(args.tables)}")
    print()

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        print("Connected to database")
    except Exception as e:
        print(f"ERROR: Cannot connect to database: {e}")
        print("\nTry setting DATABASE_URL or run from project directory with .env file")
        sys.exit(1)

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    table_configs = {
        'users': {
            'query': """
                SELECT id, email, display_name, first_name, last_name, middle_name,
                       department, job_title, is_active, is_admin, last_login_at, created_at
                FROM users ORDER BY created_at DESC
            """,
            'headers': ['ID', 'Email', 'Display Name', 'First Name', 'Last Name',
                       'Middle Name', 'Department', 'Job Title', 'Active', 'Admin',
                       'Last Login', 'Created At']
        },
        'applications': {
            'query': """
                SELECT id, name, slug, client_id, base_url, redirect_uris,
                       is_active, created_at, updated_at
                FROM applications ORDER BY name
            """,
            'headers': ['ID', 'Name', 'Slug', 'Client ID', 'Base URL',
                       'Redirect URIs', 'Active', 'Created At', 'Updated At']
        },
        'tokens': {
            'query': """
                SELECT t.id, u.email as user_email, a.name as app_name,
                       t.scopes, t.expires_at, t.created_at, t.revoked_at
                FROM oauth_tokens t
                JOIN users u ON t.user_id = u.id
                JOIN applications a ON t.application_id = a.id
                ORDER BY t.created_at DESC
                LIMIT 1000
            """,
            'headers': ['ID', 'User Email', 'Application', 'Scopes',
                       'Expires At', 'Created At', 'Revoked At']
        },
        'codes': {
            'query': """
                SELECT c.id, u.email as user_email, a.name as app_name,
                       c.redirect_uri, c.scopes, c.expires_at, c.created_at, c.used
                FROM oauth_codes c
                JOIN users u ON c.user_id = u.id
                JOIN applications a ON c.application_id = a.id
                ORDER BY c.created_at DESC
                LIMIT 1000
            """,
            'headers': ['ID', 'User Email', 'Application', 'Redirect URI',
                       'Scopes', 'Expires At', 'Created At', 'Used At']
        }
    }

    total_rows = 0
    for table in args.tables:
        config = table_configs[table]
        ws = wb.create_sheet(title=table.capitalize())

        try:
            rows = export_table(cursor, ws, table, config['query'], config['headers'])
            print(f"  {table}: {rows} rows")
            total_rows += rows
        except Exception as e:
            print(f"  {table}: ERROR - {e}")

    cursor.close()
    conn.close()

    # Save workbook
    wb.save(args.output)

    print()
    print(f"Total: {total_rows} rows exported")
    print(f"Saved to: {args.output}")
    print("=== Export complete ===")


if __name__ == '__main__':
    main()
